# Regenerates src/data/matrix.generated.ts from the real Mucho Matrix workbook
# (data/Mucho Matrix.xlsx). Run: npm run data:build
#
# Sources inside the workbook:
#   I_ActivityGroups     — one row per activity group: code, times, delegates,
#                          pickup hotel(s), mode, bus trip labels
#   T_DelegatesArrival   — one row per delegate arrival (date, pickup hotel)
#   T_DelegatesDeparture — one row per delegate departure
#
# Output matches the app's Matrix shape: hotels (with real peak in-house
# delegate counts), the ten event dates, and per hotel|date the aggregated
# movements (delegates, group rows, bus trips out/back) plus the real
# in-house count for that day (cumulative arrivals minus departures).

import json
from collections import defaultdict
from pathlib import Path

import openpyxl

HERE = Path(__file__).resolve().parent
WB = HERE.parent / "data" / "Mucho Matrix.xlsx"
OUT = HERE.parent / "src" / "data" / "matrix.generated.ts"

DATES = [f"2026-11-{d}" for d in range(22, 31)] + ["2026-12-01"]

# canonical hotels — codes, short names and addresses as the app knows them
HOTELS = [
    ("HBF", "Hilton San Diego Bayfront", "Hilton San Diego Bayfront", "1 Park Blvd"),
    ("OMN", "Omni San Diego Hotel at the Ballpark", "Omni at the Ballpark", "675 L St"),
    ("HGI", "Hilton Garden Inn San Diego Downtown/Bayside", "Hilton Garden Inn Bayside", "1655 Columbia St"),
    ("BWZ", "Best Western San Diego Zoo/Sea World", "Best Western Zoo / SeaWorld", "2575 Hotel Cir Dr"),
    ("EMB", "Embassy Suites by Hilton San Diego Bay Downtown", "Embassy Suites Bay Downtown", "601 Pacific Hwy"),
    ("SMV", "Sheraton Mission Valley San Diego Hotel", "Sheraton Mission Valley", "1433 Camino del Rio S"),
    ("COT", "Courtyard Old Town San Diego", "Courtyard Old Town", "2435 Jefferson St"),
    ("CLI", "Courtyard by Marriott Downtown Little Italy", "Courtyard Little Italy", "1112 India St"),
    ("CDT", "Courtyard by Marriott San Diego Downtown", "Courtyard Downtown", "530 Broadway"),
]
NAMES = {h[1] for h in HOTELS}

# workbook activity-code prefix -> app activity code
CODE = {"MET": "MW", "TST": "T&R"}
LABEL = {
    "CNV": "Convention Session",
    "EG": "Encouraging Gathering",
    "FS": "Field Service",
    "LEO": "Leoness Temecula Winery",
    "FFL": "Friends for Life",
    "MW": "Metropolitan Witnessing",
    "T&R": "Taste and Rejoice",
    "HBC": "Harbor Sunset Cruise",
}
GLYPH = {"CNV": "◧", "EG": "◈", "FS": "◉", "LEO": "▲", "FFL": "☰", "MW": "◍", "T&R": "◆", "HBC": "⚓"}


def mins(v):
    """A time cell (datetime.time, datetime, or 'HH:MM:SS') -> minutes since midnight."""
    if v is None or v == "":
        return None
    if hasattr(v, "hour"):
        return v.hour * 60 + v.minute
    s = str(v)
    parts = s.split(":")
    return int(parts[0]) * 60 + int(parts[1])


def iso(v):
    return str(v)[:10]


def pickups(cell):
    """A pickup cell may name one hotel or two, comma-separated."""
    out = []
    for part in str(cell or "").split(", "):
        part = part.strip()
        if part in NAMES:
            out.append(part)
    return out


wb = openpyxl.load_workbook(WB, read_only=True, data_only=True)

# ---- movements: aggregate activity-group rows -------------------------------
rows = list(wb["I_ActivityGroups"].iter_rows(values_only=True))[1:]
agg = {}
skipped = 0
for r in rows:
    name, code, s, e, date, d, pick, mode = r[0], r[1], r[2], r[3], r[4], r[5], r[6], r[7]
    bus_out, bus_back = (r[8] or ""), (r[9] or "")
    if not name or not code:
        continue
    day = iso(date)
    if day not in DATES:
        skipped += 1
        continue
    c = str(code).split("-")[0]
    c = CODE.get(c, c)
    if c not in LABEL:
        skipped += 1
        continue
    hs = pickups(pick)
    if not hs:
        skipped += 1
        continue
    sm = mins(s)
    em = mins(e)
    if sm is None:
        skipped += 1
        continue
    if em is None:
        em = min(sm + 120, 1439)
    m = "Bus" if str(mode).strip() == "Bus" else "Walking"
    d = int(float(d or 0))
    share = d // len(hs)
    for i, h in enumerate(hs):
        di = share + (d - share * len(hs) if i == 0 else 0)
        k = (h, day, c, sm, em, m)
        a = agg.setdefault(k, {"d": 0, "gr": 0, "bo": set(), "bb": set()})
        a["d"] += di
        a["gr"] += 1
        if bus_out:
            a["bo"].add(str(bus_out))
        if bus_back:
            a["bb"].add(str(bus_back))

# ---- in-house: cumulative arrivals minus departures -------------------------
arr = defaultdict(lambda: defaultdict(int))  # hotel -> iso -> count
dep = defaultdict(lambda: defaultdict(int))
for sheet, book in (("T_DelegatesArrival", arr), ("T_DelegatesDeparture", dep)):
    for r in list(wb[sheet].iter_rows(values_only=True))[1:]:
        pick, _s, date, _g, d, _acc = r[0], r[1], r[2], r[3], r[4], r[5]
        hs = pickups(pick)
        if not hs or not date:
            continue
        for h in hs:
            book[h][iso(date)] += int(float(d or 1)) // len(hs) or 1

def in_house(h, day):
    total = 0
    for d2, c2 in arr[h].items():
        if d2 <= day:
            total += c2
    for d2, c2 in dep[h].items():
        if d2 < day:  # departing today still counts this morning
            total -= c2
    return max(0, total)

# ---- assemble ---------------------------------------------------------------
by_key = {}
peak = defaultdict(int)
for _code, hname, _short, _addr in HOTELS:
    for day in DATES:
        acts = []
        for (h, d2, c, sm, em, m), a in agg.items():
            if h != hname or d2 != day:
                continue
            acts.append({
                "c": c, "n": LABEL[c], "g": GLYPH[c], "s": sm, "e": em, "m": m,
                "d": a["d"], "gr": a["gr"], "bo": len(a["bo"]), "bb": len(a["bb"]),
            })
        acts.sort(key=lambda x: (x["s"], x["c"], x["e"]))
        ih = in_house(hname, day)
        peak[hname] = max(peak[hname], ih)
        by_key[f"{hname}|{day}"] = {"acts": acts, "inHouse": ih}

hotels = [
    {"code": code, "name": name, "short": short, "addr": addr, "delegates": peak[name]}
    for code, name, short, addr in HOTELS
]
hotels.sort(key=lambda h: -h["delegates"])
matrix = {"dates": DATES, "hotels": hotels, "byKey": by_key}

n_acts = sum(len(v["acts"]) for v in by_key.values())
out = (
    "// GENERATED by scripts/build-data.py from data/Mucho Matrix.xlsx — do not edit.\n"
    f"// {len(hotels)} hotels · {len(DATES)} dates · {n_acts} movements "
    f"({len(rows)} workbook rows, {skipped} outside the window or unmapped).\n"
    "import type { Matrix } from '@/model/types'\n\n"
    f"export const MATRIX: Matrix = {json.dumps(matrix, ensure_ascii=False)} as Matrix\n"
)
OUT.write_text(out)
print(f"wrote {OUT.name}: {len(hotels)} hotels, {n_acts} movements, skipped {skipped}")
for h in hotels:
    print(f"  {h['code']}: peak in-house {h['delegates']}")
